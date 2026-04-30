from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import openpyxl

wb = Workbook()

# ── Color Palette ──────────────────────────────────────────────────────────────
NAVY      = "1B2A4A"
DARK_BLUE = "243B6E"
MID_BLUE  = "2E5FA3"
LIGHT_BLUE= "D6E4F7"
ACCENT    = "E8713A"
WHITE     = "FFFFFF"
LIGHT_GRAY= "F5F7FA"
MID_GRAY  = "D0D7E3"
DARK_GRAY = "4A4A4A"
RED_BG    = "FDECEA"
RED_FONT  = "C0392B"
AMBER_BG  = "FEF9E7"
AMBER_FONT= "B7770D"
GREEN_BG  = "E9F7EF"
GREEN_FONT= "1E8449"
CRITICAL  = "922B21"
HIGH      = "C0392B"
MODERATE  = "D4AC0D"
LOW       = "1E8449"

def hdr(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, sz=11, wrap=False, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    return c

def cell(ws, row, col, value, bg=None, fg=DARK_GRAY, bold=False, sz=10, halign="left", wrap=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if border:
        thin = Side(style="thin", color=MID_GRAY)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def merge_hdr(ws, r1, c1, r2, c2, value, bg=NAVY, fg=WHITE, bold=True, sz=12):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INSTRUCTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Instructions"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = NAVY

set_col_widths(ws1, {1:4, 2:28, 3:60, 4:28, 5:4})
for r in range(1, 60):
    set_row_height(ws1, r, 18)

# Banner
ws1.merge_cells("B1:D3")
banner = ws1["B1"]
banner.value = "AI RISK ASSESSMENT FRAMEWORK"
banner.font = Font(bold=True, color=WHITE, size=20, name="Arial")
banner.fill = PatternFill("solid", fgColor=NAVY)
banner.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws1, 1, 55)

ws1.merge_cells("B4:D4")
sub = ws1["B4"]
sub.value = "Inspired by NIST AI RMF | ISO/IEC 42001 | Enterprise GRC Best Practices"
sub.font = Font(italic=True, color=WHITE, size=10, name="Arial")
sub.fill = PatternFill("solid", fgColor=MID_BLUE)
sub.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws1, 4, 22)

# About section
merge_hdr(ws1, 6, 2, 6, 4, "ABOUT THIS TOOLKIT", bg=MID_BLUE, sz=11)
set_row_height(ws1, 6, 22)

about_rows = [
    "This toolkit provides a structured, repeatable methodology for assessing AI system risk across six critical domains.",
    "It is aligned with the NIST AI Risk Management Framework (AI RMF 1.0) and ISO/IEC 42001 AI Management Systems standard.",
    "Intended for: GRC professionals, AI governance leads, security teams, compliance officers, and portfolio demonstration.",
]
for i, txt in enumerate(about_rows, 7):
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    c = ws1.cell(row=i, column=2, value=txt)
    c.font = Font(size=10, name="Arial", color=DARK_GRAY)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    set_row_height(ws1, i, 24)

# How to use
merge_hdr(ws1, 11, 2, 11, 4, "HOW TO USE THIS WORKBOOK", bg=MID_BLUE, sz=11)
set_row_height(ws1, 11, 22)

steps = [
    ("Step 1", "Open Sheet 2 – Risk Assessment Checklist", "Answer each question for your target AI system. Score 1–5 per question."),
    ("Step 2", "Review Scoring Logic",                     "Each domain has a weight. Weighted scores roll up automatically."),
    ("Step 3", "View Sheet 3 – Scoring Dashboard",         "Dashboard auto-calculates domain scores, overall risk rating, and heat map."),
    ("Step 4", "Review Sheet 4 – Risk Summary",            "Review top risks, recommendations, and final assessment output."),
]
hdr(ws1, 12, 2, "Step", bg=DARK_BLUE, sz=10)
hdr(ws1, 12, 3, "Action", bg=DARK_BLUE, sz=10)
hdr(ws1, 12, 4, "Details", bg=DARK_BLUE, sz=10)
set_row_height(ws1, 12, 20)

for i, (step, action, detail) in enumerate(steps, 13):
    bg = LIGHT_GRAY if i % 2 else WHITE
    cell(ws1, i, 2, step, bg=MID_BLUE, fg=WHITE, bold=True, halign="center")
    cell(ws1, i, 3, action, bg=bg, bold=True)
    cell(ws1, i, 4, detail, bg=bg, wrap=True)
    set_row_height(ws1, i, 30)

# Scoring legend
merge_hdr(ws1, 18, 2, 18, 4, "SCORING SCALE REFERENCE", bg=MID_BLUE, sz=11)
set_row_height(ws1, 18, 22)

hdr(ws1, 19, 2, "Score", bg=DARK_BLUE, sz=10)
hdr(ws1, 19, 3, "Meaning", bg=DARK_BLUE, sz=10)
hdr(ws1, 19, 4, "Control Maturity", bg=DARK_BLUE, sz=10)
set_row_height(ws1, 19, 20)

score_rows = [
    ("1 – Very Low",  "Risk is negligible or fully mitigated",        "Strong, documented, tested controls in place"),
    ("2 – Low",       "Risk is minor with adequate controls",          "Controls exist and are mostly effective"),
    ("3 – Moderate",  "Risk is present with partial controls",         "Controls exist but gaps remain"),
    ("4 – High",      "Significant risk with weak/missing controls",   "Controls are minimal or ad hoc"),
    ("5 – Critical",  "Severe risk with no meaningful controls",       "No controls or controls are ineffective"),
]
colors = [GREEN_BG, GREEN_BG, AMBER_BG, RED_BG, RED_BG]
fcolors = [GREEN_FONT, GREEN_FONT, AMBER_FONT, RED_FONT, RED_FONT]
for i, (sc, mean, ctrl) in enumerate(score_rows, 20):
    cell(ws1, i, 2, sc, bg=colors[i-20], fg=fcolors[i-20], bold=True, halign="center")
    cell(ws1, i, 3, mean, bg=colors[i-20], fg=fcolors[i-20])
    cell(ws1, i, 4, ctrl, bg=colors[i-20], fg=fcolors[i-20])
    set_row_height(ws1, i, 22)

# Domain weights
merge_hdr(ws1, 26, 2, 26, 4, "DOMAIN WEIGHTS & RISK THRESHOLDS", bg=MID_BLUE, sz=11)
set_row_height(ws1, 26, 22)

hdr(ws1, 27, 2, "Domain", bg=DARK_BLUE, sz=10)
hdr(ws1, 27, 3, "Weight", bg=DARK_BLUE, sz=10)
hdr(ws1, 27, 4, "Rationale", bg=DARK_BLUE, sz=10)
set_row_height(ws1, 27, 20)

domains = [
    ("Security Risks",                      "25%", "Highest impact — adversarial attacks, data leakage, prompt injection"),
    ("Data Privacy & Protection",           "20%", "Regulatory exposure (GDPR, CCPA) and reputational damage"),
    ("Bias & Fairness",                     "15%", "Legal liability and ethical accountability"),
    ("Compliance & Legal Risk",             "15%", "Regulatory penalties and audit findings"),
    ("Accountability & Governance",         "15%", "Oversight gaps multiply all other risks"),
    ("Model Transparency & Explainability", "10%", "Operational risk; lower direct impact but enables oversight"),
]
for i, (d, w, r) in enumerate(domains, 28):
    bg = LIGHT_GRAY if i % 2 else WHITE
    cell(ws1, i, 2, d, bg=bg, bold=True)
    cell(ws1, i, 3, w, bg=bg, halign="center", bold=True, fg=MID_BLUE)
    cell(ws1, i, 4, r, bg=bg, wrap=True)
    set_row_height(ws1, i, 22)

# Footer
ws1.merge_cells("B36:D36")
ft = ws1["B36"]
ft.value = "AI Risk Assessment Framework v1.0  |  Confidential  |  For portfolio/professional use"
ft.font = Font(italic=True, color=MID_GRAY, size=9, name="Arial")
ft.alignment = Alignment(horizontal="center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RISK ASSESSMENT CHECKLIST
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Risk Assessment Checklist")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = MID_BLUE

set_col_widths(ws2, {1:4, 2:6, 3:52, 4:14, 5:14, 6:14, 7:28, 8:4})
set_row_height(ws2, 1, 50)
set_row_height(ws2, 2, 22)
set_row_height(ws2, 3, 36)
set_row_height(ws2, 4, 20)

ws2.merge_cells("B1:G1")
t = ws2["B1"]
t.value = "AI RISK ASSESSMENT — QUESTION CHECKLIST"
t.font = Font(bold=True, color=WHITE, size=16, name="Arial")
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")

ws2.merge_cells("B2:G2")
s = ws2["B2"]
s.value = "Instructions: Score each question 1 (Very Low) to 5 (Critical). Add assessor notes in the Notes column."
s.font = Font(italic=True, color=WHITE, size=10, name="Arial")
s.fill = PatternFill("solid", fgColor=MID_BLUE)
s.alignment = Alignment(horizontal="center", vertical="center")

# Column headers
cols = ["#", "Assessment Question", "Risk Level", "Score (1–5)", "Notes / Evidence"]
col_positions = [2, 3, 4, 5, 7]
for pos, label in zip(col_positions, cols):
    hdr(ws2, 3, pos, label, bg=DARK_BLUE, sz=10, wrap=True)
ws2.merge_cells(start_row=3, start_column=6, end_row=3, end_column=6)
hdr(ws2, 3, 6, "Weight", bg=DARK_BLUE, sz=10)

# ── Domain data ────────────────────────────────────────────────────────────────
DOMAINS = [
    {
        "name": "D1 — BIAS & FAIRNESS",
        "color": "5B2C6F",
        "weight": 0.15,
        "questions": [
            ("D1.1", "Has the training data been audited for demographic or historical bias?", "High"),
            ("D1.2", "Are fairness metrics (e.g., equalized odds, demographic parity) defined and monitored?", "High"),
            ("D1.3", "Does the system undergo bias testing across different demographic groups?", "High"),
            ("D1.4", "Are disparate impact analyses conducted before deployment in high-stakes decisions?", "High"),
            ("D1.5", "Is there a documented process to remediate bias when discovered post-deployment?", "Medium"),
            ("D1.6", "Are proxy variables that could introduce indirect bias identified and controlled?", "Medium"),
            ("D1.7", "Is model performance (accuracy, FPR, FNR) monitored separately per subgroup?", "Medium"),
        ],
    },
    {
        "name": "D2 — DATA PRIVACY & PROTECTION",
        "color": "1A5276",
        "weight": 0.20,
        "questions": [
            ("D2.1", "Is PII/sensitive data identified, classified, and minimized in training datasets?", "High"),
            ("D2.2", "Are data retention policies enforced, with deletion schedules documented?", "High"),
            ("D2.3", "Is consent obtained and recorded for personal data used in model training?", "High"),
            ("D2.4", "Is differential privacy or data anonymization applied to training datasets?", "High"),
            ("D2.5", "Are data processing activities documented in compliance with GDPR/CCPA requirements?", "High"),
            ("D2.6", "Is there a process to honor data subject access requests (DSAR) relating to AI outputs?", "Medium"),
            ("D2.7", "Is data shared with third-party vendors governed by DPAs and contractual controls?", "Medium"),
            ("D2.8", "Are model memorization and training data extraction risks evaluated?", "High"),
        ],
    },
    {
        "name": "D3 — MODEL TRANSPARENCY & EXPLAINABILITY",
        "color": "117A65",
        "weight": 0.10,
        "questions": [
            ("D3.1", "Can the system provide human-understandable explanations for its outputs?", "Medium"),
            ("D3.2", "Is model documentation (model cards, datasheets) maintained and publicly available?", "Medium"),
            ("D3.3", "Are feature importance and decision attribution methods implemented?", "Medium"),
            ("D3.4", "Are stakeholders (including end-users) informed that they are interacting with AI?", "High"),
            ("D3.5", "Is there a mechanism for users to contest AI-generated decisions?", "High"),
            ("D3.6", "Are model limitations, known failure modes, and confidence levels documented?", "Medium"),
            ("D3.7", "Are audit logs of model inputs and outputs retained for forensic analysis?", "Medium"),
        ],
    },
    {
        "name": "D4 — SECURITY RISKS",
        "color": "922B21",
        "weight": 0.25,
        "questions": [
            ("D4.1", "Has the model been tested against adversarial input attacks (e.g., evasion, poisoning)?", "Critical"),
            ("D4.2", "Are prompt injection and jailbreak attack vectors identified and mitigated?", "Critical"),
            ("D4.3", "Is model inversion or membership inference attack risk assessed and controlled?", "High"),
            ("D4.4", "Are API rate limits, authentication, and access controls enforced for model endpoints?", "High"),
            ("D4.5", "Is the model supply chain (training data, dependencies, libraries) verified for integrity?", "High"),
            ("D4.6", "Are outputs sanitized to prevent data leakage of sensitive training information?", "Critical"),
            ("D4.7", "Is there a documented incident response plan specific to AI system compromise?", "High"),
            ("D4.8", "Are model weights and intellectual property protected against exfiltration?", "High"),
            ("D4.9", "Is continuous monitoring in place to detect anomalous model behavior in production?", "High"),
        ],
    },
    {
        "name": "D5 — ACCOUNTABILITY & GOVERNANCE",
        "color": "784212",
        "weight": 0.15,
        "questions": [
            ("D5.1", "Is there a designated AI Risk Owner or AI Ethics Officer with clear accountability?", "High"),
            ("D5.2", "Does an AI governance committee or oversight body review high-risk AI deployments?", "High"),
            ("D5.3", "Is a human-in-the-loop (HITL) process defined for high-stakes AI decisions?", "High"),
            ("D5.4", "Are AI risk policies formally documented and reviewed at least annually?", "Medium"),
            ("D5.5", "Is an AI model inventory maintained listing all systems in production?", "Medium"),
            ("D5.6", "Are third-party AI vendors subject to risk assessment and contractual AI governance clauses?", "High"),
            ("D5.7", "Are incidents and near-misses relating to AI decisions documented and reviewed?", "Medium"),
            ("D5.8", "Is training provided to staff deploying or relying on AI system outputs?", "Medium"),
        ],
    },
    {
        "name": "D6 — COMPLIANCE & LEGAL RISK",
        "color": "1B4F72",
        "weight": 0.15,
        "questions": [
            ("D6.1", "Has a legal review been conducted for AI use cases in regulated industries or jurisdictions?", "High"),
            ("D6.2", "Is the AI system assessed against the EU AI Act risk classification requirements?", "High"),
            ("D6.3", "Are copyright/IP ownership issues resolved for training data and model outputs?", "High"),
            ("D6.4", "Are sector-specific regulations (HIPAA, FCRA, ECOA) evaluated for applicable AI use cases?", "High"),
            ("D6.5", "Is there a documented process to assess and respond to new AI regulation?", "Medium"),
            ("D6.6", "Are contractual AI liability limitations reviewed with legal counsel?", "Medium"),
            ("D6.7", "Is cross-border data transfer compliance maintained for AI training and inference?", "High"),
        ],
    },
]

ROW = 4
for domain in DOMAINS:
    # Domain header row
    set_row_height(ws2, ROW, 28)
    ws2.merge_cells(start_row=ROW, start_column=2, end_row=ROW, end_column=7)
    dh = ws2.cell(row=ROW, column=2, value=domain["name"])
    dh.font = Font(bold=True, color=WHITE, size=11, name="Arial")
    dh.fill = PatternFill("solid", fgColor=domain["color"])
    dh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ROW += 1

    for q in domain["questions"]:
        set_row_height(ws2, ROW, 32)
        bg = LIGHT_GRAY if ROW % 2 == 0 else WHITE
        # Risk level color
        rl = q[2]
        if rl == "Critical":
            rl_bg, rl_fg = "922B21", WHITE
        elif rl == "High":
            rl_bg, rl_fg = RED_BG, RED_FONT
        elif rl == "Medium":
            rl_bg, rl_fg = AMBER_BG, AMBER_FONT
        else:
            rl_bg, rl_fg = GREEN_BG, GREEN_FONT

        cell(ws2, ROW, 2, q[0], bg=bg, bold=True, halign="center", fg=MID_BLUE)
        cell(ws2, ROW, 3, q[1], bg=bg, wrap=True)
        cell(ws2, ROW, 4, rl, bg=rl_bg, fg=rl_fg, bold=True, halign="center")
        cell(ws2, ROW, 5, 3, bg=bg, halign="center", fg=MID_BLUE, bold=True)  # default score = 3
        cell(ws2, ROW, 6, domain["weight"], bg=bg, halign="center",
             fg=DARK_GRAY)
        ws2.cell(row=ROW, column=6).number_format = "0%"
        cell(ws2, ROW, 7, "", bg=bg)
        ROW += 1

    # Spacer
    set_row_height(ws2, ROW, 8)
    ROW += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SCORING DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Scoring Dashboard")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = ACCENT

set_col_widths(ws3, {1:3, 2:30, 3:14, 4:14, 5:14, 6:14, 7:18, 8:3})

# Banner
ws3.merge_cells("B1:G1")
b = ws3["B1"]
b.value = "SCORING DASHBOARD"
b.font = Font(bold=True, color=WHITE, size=16, name="Arial")
b.fill = PatternFill("solid", fgColor=NAVY)
b.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws3, 1, 45)

ws3.merge_cells("B2:G2")
s = ws3["B2"]
s.value = "Scores auto-populate from the Checklist sheet. Update individual question scores on Sheet 2."
s.font = Font(italic=True, color=WHITE, size=10, name="Arial")
s.fill = PatternFill("solid", fgColor=MID_BLUE)
s.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(ws3, 2, 20)

# Column headers
for col, label in zip([2,3,4,5,6,7],
    ["Domain", "Weight", "# Questions", "Avg Score (1–5)", "Weighted Score", "Risk Level"]):
    hdr(ws3, 4, col, label, bg=DARK_BLUE, sz=10, wrap=True)
set_row_height(ws3, 4, 30)

# Domain score rows — reference Sheet 2 score column (column E = 5)
# We need to know exact rows for each domain's questions in ws2
# Re-derive row positions
domain_row_map = {}
r = 4
for domain in DOMAINS:
    r += 1  # domain header
    start = r
    r += len(domain["questions"])
    end = r - 1
    r += 1  # spacer
    domain_row_map[domain["name"]] = (start, end)

domain_rows = {
    "D1 — BIAS & FAIRNESS": ("2. Risk Assessment Checklist", domain_row_map["D1 — BIAS & FAIRNESS"]),
    "D2 — DATA PRIVACY & PROTECTION": ("2. Risk Assessment Checklist", domain_row_map["D2 — DATA PRIVACY & PROTECTION"]),
    "D3 — MODEL TRANSPARENCY & EXPLAINABILITY": ("2. Risk Assessment Checklist", domain_row_map["D3 — MODEL TRANSPARENCY & EXPLAINABILITY"]),
    "D4 — SECURITY RISKS": ("2. Risk Assessment Checklist", domain_row_map["D4 — SECURITY RISKS"]),
    "D5 — ACCOUNTABILITY & GOVERNANCE": ("2. Risk Assessment Checklist", domain_row_map["D5 — ACCOUNTABILITY & GOVERNANCE"]),
    "D6 — COMPLIANCE & LEGAL RISK": ("2. Risk Assessment Checklist", domain_row_map["D6 — COMPLIANCE & LEGAL RISK"]),
}

weights = [0.15, 0.20, 0.10, 0.25, 0.15, 0.15]
q_counts = [7, 8, 7, 9, 8, 7]

score_row_start = 5
for i, (dname, (sheet, (sr, er))) in enumerate(domain_rows.items()):
    row = score_row_start + i
    set_row_height(ws3, row, 28)
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    safe_sheet = "'2. Risk Assessment Checklist'"
    avg_formula = f"=AVERAGE({safe_sheet}!E{sr}:E{er})"
    w = weights[i]
    weighted_formula = f"=D{row}*C{row}"

    cell(ws3, row, 2, dname, bg=bg, bold=True)
    cell(ws3, row, 3, w, bg=bg, halign="center")
    ws3.cell(row=row, column=3).number_format = "0%"
    cell(ws3, row, 4, q_counts[i], bg=bg, halign="center")
    c_avg = ws3.cell(row=row, column=5)
    c_avg.value = avg_formula
    c_avg.font = Font(size=10, name="Arial", color=MID_BLUE, bold=True)
    c_avg.fill = PatternFill("solid", fgColor=bg)
    c_avg.alignment = Alignment(horizontal="center", vertical="center")
    c_avg.number_format = "0.00"
    thin = Side(style="thin", color=MID_GRAY)
    c_avg.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c_w = ws3.cell(row=row, column=6)
    c_w.value = f"=E{row}*C{row}"
    c_w.font = Font(size=10, name="Arial", color=DARK_GRAY, bold=True)
    c_w.fill = PatternFill("solid", fgColor=bg)
    c_w.alignment = Alignment(horizontal="center", vertical="center")
    c_w.number_format = "0.00"
    c_w.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Risk level formula
    c_rl = ws3.cell(row=row, column=7)
    c_rl.value = f'=IF(E{row}>=4.5,"CRITICAL",IF(E{row}>=3.5,"HIGH",IF(E{row}>=2.5,"MODERATE","LOW")))'
    c_rl.font = Font(size=10, name="Arial", bold=True)
    c_rl.fill = PatternFill("solid", fgColor=bg)
    c_rl.alignment = Alignment(horizontal="center", vertical="center")
    c_rl.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Totals
tot_row = score_row_start + len(DOMAINS)
set_row_height(ws3, tot_row, 30)
hdr(ws3, tot_row, 2, "OVERALL WEIGHTED RISK SCORE", bg=DARK_BLUE, sz=11)
hdr(ws3, tot_row, 3, "100%", bg=DARK_BLUE, sz=11)
hdr(ws3, tot_row, 4, f"={'+'.join([str(q) for q in q_counts])}", bg=DARK_BLUE, sz=11)

c_tot = ws3.cell(row=tot_row, column=5)
c_tot.value = f"=SUM(F{score_row_start}:F{tot_row-1})"
c_tot.font = Font(bold=True, color=WHITE, size=12, name="Arial")
c_tot.fill = PatternFill("solid", fgColor=DARK_BLUE)
c_tot.alignment = Alignment(horizontal="center", vertical="center")
c_tot.number_format = "0.00"

# Merge F and G in total row
ws3.merge_cells(start_row=tot_row, start_column=6, end_row=tot_row, end_column=7)
c_rating = ws3.cell(row=tot_row, column=6)
c_rating.value = f'=IF(E{tot_row}>=4.5,"⚠ CRITICAL",IF(E{tot_row}>=3.5,"🔴 HIGH",IF(E{tot_row}>=2.5,"🟡 MODERATE","🟢 LOW")))'
c_rating.font = Font(bold=True, color=WHITE, size=12, name="Arial")
c_rating.fill = PatternFill("solid", fgColor=DARK_BLUE)
c_rating.alignment = Alignment(horizontal="center", vertical="center")

# Thresholds reference
th_row = tot_row + 2
merge_hdr(ws3, th_row, 2, th_row, 7, "RISK RATING THRESHOLDS", bg=MID_BLUE, sz=11)
set_row_height(ws3, th_row, 22)

thresholds = [
    ("LOW",      "Score < 2.5",  "Acceptable risk. Maintain controls and monitor.", GREEN_BG, GREEN_FONT),
    ("MODERATE", "2.5 – 3.4",   "Elevated risk. Implement mitigations within 90 days.", AMBER_BG, AMBER_FONT),
    ("HIGH",     "3.5 – 4.4",   "Significant risk. Immediate remediation plan required.", RED_BG, RED_FONT),
    ("CRITICAL", "Score ≥ 4.5", "Severe risk. Consider halting deployment pending remediation.", "FDECEA", CRITICAL),
]
for j, (r, sc, desc, bg, fg) in enumerate(thresholds, th_row + 1):
    set_row_height(ws3, j, 24)
    cell(ws3, j, 2, r, bg=bg, fg=fg, bold=True, halign="center")
    cell(ws3, j, 3, sc, bg=bg, fg=fg, bold=True, halign="center")
    ws3.merge_cells(start_row=j, start_column=4, end_row=j, end_column=7)
    cell(ws3, j, 4, desc, bg=bg, fg=fg)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — RISK SUMMARY (ChatGPT Sample Assessment)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Risk Summary")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = "C0392B"

set_col_widths(ws4, {1:3, 2:28, 3:16, 4:16, 5:22, 6:30, 7:3})
set_row_height(ws4, 1, 50)

ws4.merge_cells("B1:F1")
b = ws4["B1"]
b.value = "AI RISK ASSESSMENT REPORT — SAMPLE: ChatGPT (GPT-4)"
b.font = Font(bold=True, color=WHITE, size=15, name="Arial")
b.fill = PatternFill("solid", fgColor=NAVY)
b.alignment = Alignment(horizontal="center", vertical="center")

meta = [
    ("System Assessed:", "ChatGPT (GPT-4, OpenAI)"),
    ("Assessment Date:", "April 2026"),
    ("Assessor:",        "AI GRC Portfolio Project"),
    ("Classification:",  "CONFIDENTIAL — SAMPLE"),
    ("Framework:",       "NIST AI RMF 1.0 | ISO/IEC 42001"),
]
for i, (k, v) in enumerate(meta, 2):
    set_row_height(ws4, i, 20)
    cell(ws4, i, 2, k, bg=LIGHT_BLUE, bold=True, fg=DARK_BLUE)
    ws4.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
    cell(ws4, i, 3, v, bg=LIGHT_BLUE, fg=DARK_GRAY)

# Domain scores for ChatGPT sample
merge_hdr(ws4, 8, 2, 8, 6, "DOMAIN RISK SCORES — CHATGPT SAMPLE ASSESSMENT", bg=MID_BLUE, sz=11)
set_row_height(ws4, 8, 22)

for col, label in zip([2,3,4,5,6], ["Domain", "Weight", "Score (1–5)", "Weighted Score", "Risk Level"]):
    hdr(ws4, 9, col, label, bg=DARK_BLUE, sz=10)
set_row_height(ws4, 9, 20)

chatgpt_scores = [
    ("Bias & Fairness",                       0.15, 3.4, "MODERATE"),
    ("Data Privacy & Protection",             0.20, 3.8, "HIGH"),
    ("Model Transparency & Explainability",   0.10, 3.6, "HIGH"),
    ("Security Risks",                        0.25, 4.1, "HIGH"),
    ("Accountability & Governance",           0.15, 3.2, "MODERATE"),
    ("Compliance & Legal Risk",               0.15, 3.5, "HIGH"),
]
overall = sum(s * w for _, w, s, _ in chatgpt_scores)

for i, (d, w, sc, rl) in enumerate(chatgpt_scores, 10):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    if rl == "CRITICAL": rl_bg, rl_fg = "922B21", WHITE
    elif rl == "HIGH":   rl_bg, rl_fg = RED_BG, RED_FONT
    elif rl == "MODERATE": rl_bg, rl_fg = AMBER_BG, AMBER_FONT
    else: rl_bg, rl_fg = GREEN_BG, GREEN_FONT

    set_row_height(ws4, i, 24)
    cell(ws4, i, 2, d, bg=bg, bold=True)
    cell(ws4, i, 3, w, bg=bg, halign="center")
    ws4.cell(row=i, column=3).number_format = "0%"
    cell(ws4, i, 4, sc, bg=bg, halign="center", bold=True, fg=MID_BLUE)
    ws_col = ws4.cell(row=i, column=5)
    ws_col.value = round(sc * w, 3)
    ws_col.number_format = "0.000"
    ws_col.font = Font(size=10, name="Arial")
    ws_col.fill = PatternFill("solid", fgColor=bg)
    ws_col.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color=MID_GRAY)
    ws_col.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell(ws4, i, 6, rl, bg=rl_bg, fg=rl_fg, bold=True, halign="center")

# Overall score row
set_row_height(ws4, 16, 30)
hdr(ws4, 16, 2, "OVERALL RISK SCORE", bg=DARK_BLUE, sz=11)
hdr(ws4, 16, 3, "100%", bg=DARK_BLUE, sz=11)
c_os = ws4.cell(row=16, column=4, value=round(overall, 2))
c_os.font = Font(bold=True, color=WHITE, size=12, name="Arial")
c_os.fill = PatternFill("solid", fgColor=DARK_BLUE)
c_os.alignment = Alignment(horizontal="center", vertical="center")
ws4.merge_cells(start_row=16, start_column=5, end_row=16, end_column=6)
hdr(ws4, 16, 5, "🔴  HIGH RISK (3.72)", bg=RED_FONT, sz=12)

# Top 5 Risks
merge_hdr(ws4, 18, 2, 18, 6, "TOP 5 IDENTIFIED RISKS", bg=MID_BLUE, sz=11)
set_row_height(ws4, 18, 22)

for col, lbl in zip([2,3,4,5,6], ["#", "Risk", "Domain", "Severity", "Recommended Mitigation"]):
    hdr(ws4, 19, col, lbl, bg=DARK_BLUE, sz=10)
set_row_height(ws4, 19, 20)

top_risks = [
    ("1", "Prompt injection enabling policy bypass or data exfiltration",
     "Security", "CRITICAL", "Implement input sanitization, output filtering, and adversarial red-teaming"),
    ("2", "Training data memorization leaking PII in model outputs",
     "Data Privacy", "HIGH", "Apply differential privacy, output monitoring, and PII detection filters"),
    ("3", "Opaque reasoning with no contestation mechanism for high-stakes outputs",
     "Transparency", "HIGH", "Deploy model cards, confidence scores, and user appeal workflows"),
    ("4", "Cross-border data transfer non-compliance under GDPR/EU AI Act",
     "Compliance", "HIGH", "Legal review of data flows; SCCs or data localization controls"),
    ("5", "Demographic bias in content moderation or decision-support outputs",
     "Bias & Fairness", "MODERATE", "Regular fairness audits, subgroup performance monitoring, bias red-teams"),
]
sev_colors = {"CRITICAL": ("922B21", WHITE), "HIGH": (RED_BG, RED_FONT), "MODERATE": (AMBER_BG, AMBER_FONT)}

for i, (num, risk, domain, sev, mit) in enumerate(top_risks, 20):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    sc, fc = sev_colors.get(sev, (GREEN_BG, GREEN_FONT))
    set_row_height(ws4, i, 36)
    cell(ws4, i, 2, num, bg=MID_BLUE, fg=WHITE, bold=True, halign="center")
    cell(ws4, i, 3, risk, bg=bg, wrap=True)
    cell(ws4, i, 4, domain, bg=bg, bold=True, fg=MID_BLUE)
    cell(ws4, i, 5, sev, bg=sc, fg=fc, bold=True, halign="center")
    cell(ws4, i, 6, mit, bg=bg, wrap=True)

# Recommendations
merge_hdr(ws4, 26, 2, 26, 6, "KEY RECOMMENDATIONS", bg=MID_BLUE, sz=11)
set_row_height(ws4, 26, 22)

recs = [
    ("Immediate (0–30 days)",  "Conduct adversarial red-team exercise targeting prompt injection; patch output filters."),
    ("Short-term (30–90 days)", "Commission independent bias audit; implement subgroup monitoring dashboards."),
    ("Medium-term (90–180 days)", "Publish model card; establish AI governance committee and AI risk register."),
    ("Long-term (180+ days)",  "Achieve ISO/IEC 42001 certification; integrate AI risk into enterprise ERM program."),
]
for j, (timeframe, rec) in enumerate(recs, 27):
    bg = AMBER_BG if j % 2 == 0 else WHITE
    set_row_height(ws4, j, 30)
    cell(ws4, j, 2, timeframe, bg=bg, bold=True, fg=AMBER_FONT)
    ws4.merge_cells(start_row=j, start_column=3, end_row=j, end_column=6)
    cell(ws4, j, 3, rec, bg=bg, wrap=True)

# Footer
ws4.merge_cells("B32:F32")
f = ws4["B32"]
f.value = "AI Risk Assessment Framework v1.0  |  Sample Assessment: ChatGPT (GPT-4)  |  Not for production use without customization"
f.font = Font(italic=True, size=9, color=MID_GRAY, name="Arial")
f.alignment = Alignment(horizontal="center")

# ── Save ───────────────────────────────────────────────────────────────────────
out = "/home/claude/AI_Risk_Assessment_Toolkit.xlsx"
wb.save(out)
print("Saved:", out)
